import pytest
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from login_page import LoginPage
from openpyxl import load_workbook
from datetime import datetime

EXCEL_PATH = "login_test_data.xlsx"
TESTER_NAME = "Tester1"
LOGIN_URL = "https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"

def read_test_data():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        test_id, username, password, *_ = row
        data.append((test_id, username, password))
    return data

@pytest.fixture(scope="module")
def driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--remote-allow-origins=*")
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)
    driver.maximize_window()
    yield driver
    driver.quit()

@pytest.mark.parametrize("test_id, username, password", read_test_data())
def test_login(driver, test_id, username, password):
    driver.get(LOGIN_URL)
    login_page = LoginPage(driver)
    login_page.enter_username(username)
    login_page.enter_password(password)
    login_page.click_login()

    result = "Pass" if login_page.is_login_successful() else "Fail"

    # Update Excel file with results
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(test_id):
            row[3].value = datetime.now().strftime("%Y-%m-%d")   # Date
            row[4].value = datetime.now().strftime("%H:%M:%S")   # Time of Test
            row[5].value = TESTER_NAME                            # Name of Tester
            row[6].value = result                                 # Test Result
            break
    wb.save(EXCEL_PATH)

    assert result == "Pass", f"Login failed for user: {username}"
